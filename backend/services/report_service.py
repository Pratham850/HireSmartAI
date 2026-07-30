import os
import json
import csv
from datetime import datetime
from typing import Optional, List, Dict, Any
from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from models.candidate import Candidate
from models.analysis_result import AnalysisResult
from models.job_description import JobDescription
from schemas.report import ReportResponse
from config import settings

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def _fetch_report_data(self, jd_id: Optional[int] = None) -> List[Dict[str, Any]]:
        query = select(AnalysisResult, Candidate, JobDescription).join(
            Candidate, AnalysisResult.candidate_id == Candidate.id
        ).join(
            JobDescription, AnalysisResult.job_description_id == JobDescription.id
        )

        if jd_id:
            query = query.where(AnalysisResult.job_description_id == jd_id)

        query = query.order_by(AnalysisResult.match_score.desc())
        result = await self.db.execute(query)
        rows = result.all()

        data_rows = []
        for ar, cand, jd in rows:
            data_rows.append({
                "candidate_id": cand.id,
                "name": cand.name,
                "email": cand.email or "N/A",
                "phone": cand.phone or "N/A",
                "job_role": jd.role,
                "match_score": ar.match_score,
                "recommendation": ar.recommendation,
                "skills": ", ".join(cand.skills or []),
                "missing_skills": ", ".join(ar.missing_skills or []),
                "strengths": ", ".join(ar.strengths or []),
                "weaknesses": ", ".join(ar.weaknesses or []),
                "reason": ar.reason or ""
            })

        if not data_rows:
            cand_stmt = select(Candidate)
            cands = (await self.db.execute(cand_stmt)).scalars().all()
            for cand in cands:
                data_rows.append({
                    "candidate_id": cand.id,
                    "name": cand.name,
                    "email": cand.email or "N/A",
                    "phone": cand.phone or "N/A",
                    "job_role": "Unspecified",
                    "match_score": 0.0,
                    "recommendation": "Pending Analysis",
                    "skills": ", ".join(cand.skills or []),
                    "missing_skills": "N/A",
                    "strengths": "N/A",
                    "weaknesses": "N/A",
                    "reason": "Not analyzed yet"
                })
        return data_rows

    async def generate_excel_report(self, jd_id: Optional[int] = None) -> ReportResponse:
        data = await self._fetch_report_data(jd_id)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if not openpyxl:
            # Fallback to CSV format if openpyxl is not installed
            return await self.generate_csv_report(jd_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Rankings"

        headers = [
            "Candidate ID", "Full Name", "Email", "Phone", "Target Role",
            "Match Score (%)", "Recommendation", "Skills", "Missing Skills", "Reason"
        ]
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for row_idx, row in enumerate(data, start=2):
            ws.append([
                row["candidate_id"],
                row["name"],
                row["email"],
                row["phone"],
                row["job_role"],
                row["match_score"],
                row["recommendation"],
                row["skills"],
                row["missing_skills"],
                row["reason"]
            ])

            rec_cell = ws.cell(row=row_idx, column=7)
            if row["recommendation"] == "Shortlist":
                rec_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                rec_cell.font = Font(color="006100", bold=True)
            elif row["recommendation"] == "Reject":
                rec_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                rec_cell.font = Font(color="9C0006", bold=True)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        filename = f"hiresmart_report_{timestamp}.xlsx"
        filepath = os.path.join(settings.REPORT_DIR, filename)
        wb.save(filepath)

        return ReportResponse(
            message="Excel report successfully generated",
            file_path=filepath,
            download_url=f"/reports/download/{filename}",
            file_type="excel"
        )

    async def generate_csv_report(self, jd_id: Optional[int] = None) -> ReportResponse:
        data = await self._fetch_report_data(jd_id)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"hiresmart_report_{timestamp}.csv"
        filepath = os.path.join(settings.REPORT_DIR, filename)

        fieldnames = [
            "candidate_id", "name", "email", "phone", "job_role",
            "match_score", "recommendation", "skills", "missing_skills",
            "strengths", "weaknesses", "reason"
        ]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in data:
                writer.writerow(row)

        return ReportResponse(
            message="CSV report successfully generated",
            file_path=filepath,
            download_url=f"/reports/download/{filename}",
            file_type="csv"
        )

    async def generate_json_report(self, jd_id: Optional[int] = None) -> ReportResponse:
        data = await self._fetch_report_data(jd_id)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"hiresmart_report_{timestamp}.json"
        filepath = os.path.join(settings.REPORT_DIR, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({"generated_at": datetime.now().isoformat(), "candidates": data}, f, indent=2)

        return ReportResponse(
            message="JSON report successfully generated",
            file_path=filepath,
            download_url=f"/reports/download/{filename}",
            file_type="json"
        )
